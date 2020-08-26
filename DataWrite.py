# -*- coding: utf-8 -*-
"""
Created on Wed Jan 22 04:23:54 2020

@author: 13052
"""
import openpyxl
import numpy as np
import pandas as pd

re = pd.read_table('temdata.txt',header = None, delimiter='\t')
re2= pd.read_table('result_t.txt',header = None, delimiter='\t')
dis = pd.read_table('distance.txt',header = None, delimiter='\t')
res = re.values
res2 = re2.values
distance = dis.values


data = openpyxl.open('res.xlsx')
sheet = data['Sheet1']
points = ["I",
"E",
"L",
"K",
"A",
"J",
"O",
"F",
"P",
"N",
"M",
"B",
"D",
"G",
"C",
"H"];
          
length = [0]*111          
for i in range(0,111):
    for j in range(3,0,-1):
        if res2[i,j] != 0:
            length[i]=j
            break
                  

kk = 1;
for i in range(0,111):
    for j in range(0,4): 
        if res2[i,j] != 0 :
            sheet.cell(row=kk,column=1,value= i+1)
            if j!=0 and res[res2[i,j]-1,0] != res[res2[i,j-1]-1,1]:
                sheet.cell(row=kk,column=2,value= '空驶')
                sheet.cell(row=kk,column=3,value= points[res[res2[i,j-1]-1,1]-1])
                sheet.cell(row=kk,column=4,value= points[res[res2[i,j]-1,0]-1])
                sheet.cell(row=kk,column=5,value= res[res2[i,j-1]-1,3])
                sheet.cell(row=kk,column=6,value= res[res2[i,j-1]-1,3]+distance[res[res2[i,j-1]-1,1]-1,res[res2[i,j]-1,0]-1])
                
                
                kk=kk+1
            sheet.cell(row=kk,column=1,value= i+1)
            sheet.cell(row=kk,column=2,value= res2[i,j])
            for w in range(3,5):
                sheet.cell(row=kk,column=w,value= points[res[res2[i,j]-1,w-3]-1]) 
            for w in range(5,7):
                sheet.cell(row=kk,column=w,value= res[res2[i,j]-1,w-3]) 
            kk=kk+1
            if j == length[i] and res[res2[i,j]-1,1] != res[res2[i,0]-1,0]:
                sheet.cell(row=kk,column=1,value= i+1)
                sheet.cell(row=kk,column=2,value='空驶')
                sheet.cell(row=kk,column=3,value= points[res[res2[i,j]-1,1]-1])
                sheet.cell(row=kk,column=4,value= points[res[res2[i,0]-1,0]-1])
                sheet.cell(row=kk,column=5,value= res[res2[i,j]-1,3])
                sheet.cell(row=kk,column=6,value= res[res2[i,j]-1,3]+distance[res[res2[i,j]-1,1]-1,res[res2[i,0]-1,0]-1])
                kk+=1

    data.save('res.xlsx')   
    print('----Finish writing----')